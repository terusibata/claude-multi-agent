# -*- coding: utf-8 -*-
"""
Excelファイル用ツール

AIエージェントがExcelファイルを理解するための軽量ツール。
3つの機能を提供:
1. get_sheet_info: シート一覧と基本情報を取得
2. get_sheet_csv: 指定シートの内容をCSV Markdown形式で取得（行範囲指定可能）
3. search_workbook: ワークブック全体からキーワード検索
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from io import StringIO
from typing import TYPE_CHECKING, Any, Optional, TypedDict

import structlog

if TYPE_CHECKING:
    from app.services.workspace_service import WorkspaceService

logger = structlog.get_logger(__name__)


# =============================================================================
# Type Definitions
# =============================================================================

class SheetInfo(TypedDict):
    """シート情報の型定義"""
    name: str
    rows: int
    cols: int
    range: str
    has_print_area: bool


class WorkbookInfo(TypedDict):
    """ワークブック情報の型定義"""
    filename: str
    sheet_count: int
    sheets: list[SheetInfo]


class SheetCSVResult(TypedDict):
    """シートCSV取得結果の型定義"""
    sheet_name: str
    range: str
    total_rows: int
    total_cols: int
    returned_rows: int
    start_row: int
    end_row: int
    has_more: bool
    csv_markdown: str


class SearchHit(TypedDict):
    """検索ヒットの型定義"""
    sheet: str
    cell: str
    row: int
    col: int
    value: str
    context: str


class SearchResult(TypedDict):
    """検索結果の型定義"""
    query: str
    total_hits: int
    hits: list[SearchHit]


# =============================================================================
# Constants
# =============================================================================

DEFAULT_MAX_ROWS = 100  # デフォルトの最大取得行数


# =============================================================================
# Internal Utilities
# =============================================================================

def _normalize_text(text: str) -> str:
    """テキストの正規化（Unicode NFC、制御文字除去）"""
    if not text:
        return ""

    # Unicode NFC正規化
    text = unicodedata.normalize("NFC", text)

    # 制御文字を除去（改行・タブは保持）
    result = []
    for ch in text:
        code = ord(ch)
        if ch in ('\n', '\r', '\t'):
            result.append(ch)
        elif code < 0x20 or code == 0x7F or (0x80 <= code <= 0x9F):
            continue
        elif code in (0x200B, 0x200C, 0x200D, 0x2060, 0xFEFF):
            continue
        else:
            result.append(ch)

    return "".join(result)


def _get_cell_value(cell) -> str:
    """セルの表示値を取得"""
    value = cell.value

    if value is None:
        return ""

    if hasattr(cell, 'is_date') and cell.is_date:
        try:
            import datetime
            if isinstance(value, (datetime.datetime, datetime.date)):
                return value.strftime("%Y-%m-%d")
        except Exception:
            pass

    text = str(value)
    text = _normalize_text(text)

    return text.strip()


def _get_print_area(ws) -> Optional[tuple[int, int, int, int]]:
    """印刷領域を取得（min_row, min_col, max_row, max_col）"""
    try:
        from openpyxl.utils import range_boundaries

        pa = ws.print_area
        if pa:
            pa_str = pa if isinstance(pa, str) else str(pa[0]) if pa else None
            if pa_str:
                if '!' in pa_str:
                    pa_str = pa_str.split('!')[-1]

                min_col, min_row, max_col, max_row = range_boundaries(pa_str)
                return (min_row, min_col, max_row, max_col)
    except Exception:
        pass

    return None


def _get_used_range(ws) -> tuple[int, int, int, int]:
    """使用範囲を取得（min_row, min_col, max_row, max_col）"""
    min_row = ws.min_row or 1
    min_col = ws.min_column or 1
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    return (min_row, min_col, max_row, max_col)


def _build_merged_lookup(ws, area: tuple[int, int, int, int]) -> dict[tuple[int, int], tuple[int, int]]:
    """結合セルのルックアップを構築（各セル -> 左上セル）"""
    lookup: dict[tuple[int, int], tuple[int, int]] = {}
    area_min_row, area_min_col, area_max_row, area_max_col = area

    for merged_range in getattr(ws, "merged_cells", []):
        try:
            r0, c0 = merged_range.min_row, merged_range.min_col
            r1, c1 = merged_range.max_row, merged_range.max_col

            if r1 < area_min_row or r0 > area_max_row:
                continue
            if c1 < area_min_col or c0 > area_max_col:
                continue
            if r0 < area_min_row or c0 < area_min_col:
                continue

            for r in range(r0, r1 + 1):
                for c in range(c0, c1 + 1):
                    if area_min_row <= r <= area_max_row and area_min_col <= c <= area_max_col:
                        lookup[(r, c)] = (r0, c0)
        except Exception:
            continue

    return lookup


def _coords_to_range_string(min_row: int, min_col: int, max_row: int, max_col: int) -> str:
    """座標をExcel範囲文字列に変換（例: A1:D10）"""
    from openpyxl.utils import get_column_letter

    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    return f"{start}:{end}"


def _load_workbook_from_bytes(content: bytes):
    """バイトデータからワークブックを読み込む"""
    import openpyxl

    return openpyxl.load_workbook(
        filename=io.BytesIO(content),
        read_only=False,
        data_only=True
    )


def _check_old_format(file_path: str) -> dict[str, Any] | None:
    """古いOffice形式（.xls）のチェック"""
    if file_path.lower().endswith(".xls"):
        return {
            "content": [{
                "type": "text",
                "text": (
                    f"エラー: '{file_path}' は古いExcel形式（.xls）です。\n\n"
                    "このツールは .xlsx/.xlsm（Office Open XML）形式のみ対応しています。\n"
                    ".xls（BIFF形式）ファイルは openpyxl では読み取れません。\n\n"
                    "対処方法:\n"
                    "1. Microsoft Excel で .xlsx 形式に変換して再アップロード\n"
                    "2. LibreOffice Calc で .xlsx 形式に変換して再アップロード\n"
                    "3. オンライン変換ツールを使用"
                ),
            }],
            "is_error": True,
        }
    return None


# =============================================================================
# Core Functions
# =============================================================================

def get_sheet_info(content: bytes, filename: str) -> WorkbookInfo:
    """
    Excelファイルのシート情報を取得する。

    Args:
        content: Excelファイルのバイトデータ
        filename: ファイル名

    Returns:
        WorkbookInfo: ワークブック情報
    """
    wb = _load_workbook_from_bytes(content)

    sheets: list[SheetInfo] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print_area = _get_print_area(ws)
        has_print_area = print_area is not None

        if print_area:
            min_row, min_col, max_row, max_col = print_area
        else:
            min_row, min_col, max_row, max_col = _get_used_range(ws)

        rows = max_row - min_row + 1
        cols = max_col - min_col + 1
        range_str = _coords_to_range_string(min_row, min_col, max_row, max_col)

        sheets.append(SheetInfo(
            name=sheet_name,
            rows=rows,
            cols=cols,
            range=range_str,
            has_print_area=has_print_area
        ))

    wb.close()

    return WorkbookInfo(
        filename=filename,
        sheet_count=len(sheets),
        sheets=sheets
    )


def get_sheet_csv(
    content: bytes,
    sheet_name: str,
    *,
    start_row: int | None = None,
    end_row: int | None = None,
    max_rows: int = DEFAULT_MAX_ROWS,
    use_print_area: bool = True
) -> SheetCSVResult:
    """
    指定シートの内容をCSV Markdown形式で取得する。

    Args:
        content: Excelファイルのバイトデータ
        sheet_name: 取得するシート名
        start_row: 開始行（1始まり、Noneの場合は先頭から）
        end_row: 終了行（1始まり、Noneの場合はmax_rowsまで）
        max_rows: 最大取得行数（デフォルト: 100）。end_row指定時は無視
        use_print_area: 印刷領域を使用するか（デフォルト: True）

    Returns:
        SheetCSVResult: シートCSV取得結果
    """
    wb = _load_workbook_from_bytes(content)

    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"シートが見つかりません: {sheet_name}")

    ws = wb[sheet_name]

    # 対象範囲を決定
    print_area = _get_print_area(ws) if use_print_area else None

    if print_area:
        area_min_row, area_min_col, area_max_row, area_max_col = print_area
    else:
        area_min_row, area_min_col, area_max_row, area_max_col = _get_used_range(ws)

    total_rows = area_max_row - area_min_row + 1
    total_cols = area_max_col - area_min_col + 1

    # 実際の取得範囲を決定
    actual_start_row = start_row if start_row is not None else area_min_row

    if end_row is not None:
        actual_end_row = min(end_row, area_max_row)
    else:
        actual_end_row = min(actual_start_row + max_rows - 1, area_max_row)

    # 範囲チェック
    actual_start_row = max(actual_start_row, area_min_row)
    actual_end_row = min(actual_end_row, area_max_row)

    area = (actual_start_row, area_min_col, actual_end_row, area_max_col)

    # 結合セルのルックアップを構築
    merged_lookup = _build_merged_lookup(ws, area)

    # CSVデータを抽出
    csv_rows: list[list[str]] = []

    for r in range(actual_start_row, actual_end_row + 1):
        row_values: list[str] = []

        for c in range(area_min_col, area_max_col + 1):
            cell = ws.cell(row=r, column=c)

            top_left = merged_lookup.get((r, c))
            if top_left:
                if (r, c) == top_left:
                    text = _get_cell_value(cell)
                else:
                    text = ""
            else:
                text = _get_cell_value(cell)

            text = text.replace('\r\n', ' ').replace('\n', ' ').replace('\r', ' ')
            row_values.append(text)

        csv_rows.append(row_values)

    wb.close()

    # CSV文字列を生成（RFC 4180準拠）
    output = StringIO()
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL, lineterminator='\n')
    writer.writerows(csv_rows)
    csv_content = output.getvalue()

    csv_markdown = f"```csv\n{csv_content}```"

    has_more = actual_end_row < area_max_row

    return SheetCSVResult(
        sheet_name=sheet_name,
        range=_coords_to_range_string(actual_start_row, area_min_col, actual_end_row, area_max_col),
        total_rows=total_rows,
        total_cols=total_cols,
        returned_rows=len(csv_rows),
        start_row=actual_start_row,
        end_row=actual_end_row,
        has_more=has_more,
        csv_markdown=csv_markdown
    )


def search_workbook(
    content: bytes,
    query: str,
    *,
    case_sensitive: bool = False,
    max_hits: int = 50
) -> SearchResult:
    """
    ワークブック全体からキーワード検索を行う。

    Args:
        content: Excelファイルのバイトデータ
        query: 検索キーワード
        case_sensitive: 大文字小文字を区別するか（デフォルト: False）
        max_hits: 最大ヒット数（デフォルト: 50）

    Returns:
        SearchResult: 検索結果
    """
    from openpyxl.utils import get_column_letter

    wb = _load_workbook_from_bytes(content)

    hits: list[SearchHit] = []

    # 検索パターンを準備
    flags = 0 if case_sensitive else re.IGNORECASE
    try:
        pattern = re.compile(re.escape(query), flags)
    except re.error:
        wb.close()
        raise ValueError(f"無効な検索クエリ: {query}")

    for sheet_name in wb.sheetnames:
        if len(hits) >= max_hits:
            break

        ws = wb[sheet_name]
        min_row, min_col, max_row, max_col = _get_used_range(ws)

        for r in range(min_row, max_row + 1):
            if len(hits) >= max_hits:
                break

            # 行のセル値をキャッシュ（コンテキスト生成用）
            row_values: dict[int, str] = {}
            for c in range(min_col, max_col + 1):
                cell = ws.cell(row=r, column=c)
                row_values[c] = _get_cell_value(cell)

            for c in range(min_col, max_col + 1):
                if len(hits) >= max_hits:
                    break

                value = row_values[c]
                if not value:
                    continue

                if pattern.search(value):
                    # コンテキストを生成（前後2セル）
                    context_parts = []
                    for ctx_col in range(max(min_col, c - 2), min(max_col + 1, c + 3)):
                        ctx_value = row_values.get(ctx_col, "")
                        if ctx_value:
                            col_letter = get_column_letter(ctx_col)
                            if ctx_col == c:
                                context_parts.append(f"[{col_letter}:{ctx_value}]")
                            else:
                                context_parts.append(f"{col_letter}:{ctx_value}")

                    context = " | ".join(context_parts)

                    hits.append(SearchHit(
                        sheet=sheet_name,
                        cell=f"{get_column_letter(c)}{r}",
                        row=r,
                        col=c,
                        value=value,
                        context=context
                    ))

    wb.close()

    return SearchResult(
        query=query,
        total_hits=len(hits),
        hits=hits
    )


# =============================================================================
# Tool Handlers
# =============================================================================

async def get_sheet_info_handler(
    workspace_service: "WorkspaceService",
    tenant_id: str,
    conversation_id: str,
    args: dict[str, Any],
) -> dict[str, Any]:
    """
    Excelファイルのシート情報を取得するハンドラー

    Args:
        args:
            file_path: ファイルパス
    """
    file_path = args.get("file_path", "")

    # 古い形式チェック
    old_format_error = _check_old_format(file_path)
    if old_format_error:
        return old_format_error

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return {
            "content": [{"type": "text", "text": "エラー: openpyxlライブラリがインストールされていません。"}],
            "is_error": True,
        }

    try:
        content, filename, _ = await workspace_service.download_file(
            tenant_id, conversation_id, file_path
        )

        info = get_sheet_info(content, filename)

        # 結果をテキスト形式でフォーマット
        result_lines = [
            f"# Excel情報: {info['filename']}",
            f"シート数: {info['sheet_count']}",
            "",
            "## シート一覧",
        ]

        for sheet in info['sheets']:
            result_lines.append(f"")
            result_lines.append(f"### {sheet['name']}")
            result_lines.append(f"- 範囲: {sheet['range']} ({sheet['rows']}行 x {sheet['cols']}列)")
            if sheet['has_print_area']:
                result_lines.append(f"- 印刷領域: 設定済み")

        result_lines.append("")
        result_lines.append("---")
        result_lines.append("データを取得するには `get_sheet_csv` を使用してください。")
        result_lines.append("キーワード検索には `search_workbook` を使用してください。")

        return {
            "content": [{"type": "text", "text": "\n".join(result_lines)}],
        }
    except FileNotFoundError:
        return {
            "content": [{"type": "text", "text": f"ファイルが見つかりません: {file_path}"}],
            "is_error": True,
        }
    except Exception as e:
        logger.error("Excel情報取得エラー", error=str(e), file_path=file_path)
        return {
            "content": [{"type": "text", "text": f"読み込みエラー: {str(e)}"}],
            "is_error": True,
        }


async def get_sheet_csv_handler(
    workspace_service: "WorkspaceService",
    tenant_id: str,
    conversation_id: str,
    args: dict[str, Any],
) -> dict[str, Any]:
    """
    指定シートの内容をCSV Markdown形式で取得するハンドラー

    Args:
        args:
            file_path: ファイルパス
            sheet_name: シート名
            start_row: 開始行（1始まり、省略時は先頭から）
            end_row: 終了行（1始まり、省略時はmax_rowsまで）
            max_rows: 最大取得行数（デフォルト: 100）
            use_print_area: 印刷領域を使用するか（デフォルト: true）
    """
    file_path = args.get("file_path", "")
    sheet_name = args.get("sheet_name", "")
    start_row = args.get("start_row")
    end_row = args.get("end_row")
    max_rows = args.get("max_rows", DEFAULT_MAX_ROWS)
    use_print_area = args.get("use_print_area", True)

    # 古い形式チェック
    old_format_error = _check_old_format(file_path)
    if old_format_error:
        return old_format_error

    if not sheet_name:
        return {
            "content": [{"type": "text", "text": "エラー: sheet_name（シート名）を指定してください。\nget_sheet_info でシート一覧を確認できます。"}],
            "is_error": True,
        }

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return {
            "content": [{"type": "text", "text": "エラー: openpyxlライブラリがインストールされていません。"}],
            "is_error": True,
        }

    try:
        content, _, _ = await workspace_service.download_file(
            tenant_id, conversation_id, file_path
        )

        result = get_sheet_csv(
            content,
            sheet_name,
            start_row=start_row,
            end_row=end_row,
            max_rows=max_rows,
            use_print_area=use_print_area,
        )

        # 結果をテキスト形式でフォーマット
        result_lines = [
            f"# {result['sheet_name']}",
            f"範囲: {result['range']}",
            f"サイズ: {result['returned_rows']}/{result['total_rows']}行 x {result['total_cols']}列",
            "",
            result['csv_markdown'],
        ]

        if result['has_more']:
            result_lines.append("")
            result_lines.append(f"---")
            result_lines.append(f"まだ続きがあります。次を取得するには:")
            result_lines.append(f"`start_row={result['end_row'] + 1}` を指定してください。")

        return {
            "content": [{"type": "text", "text": "\n".join(result_lines)}],
        }
    except FileNotFoundError:
        return {
            "content": [{"type": "text", "text": f"ファイルが見つかりません: {file_path}"}],
            "is_error": True,
        }
    except ValueError as e:
        return {
            "content": [{"type": "text", "text": str(e)}],
            "is_error": True,
        }
    except Exception as e:
        logger.error("ExcelCSV取得エラー", error=str(e), file_path=file_path)
        return {
            "content": [{"type": "text", "text": f"読み込みエラー: {str(e)}"}],
            "is_error": True,
        }


async def search_workbook_handler(
    workspace_service: "WorkspaceService",
    tenant_id: str,
    conversation_id: str,
    args: dict[str, Any],
) -> dict[str, Any]:
    """
    ワークブック全体からキーワード検索を行うハンドラー

    Args:
        args:
            file_path: ファイルパス
            query: 検索キーワード
            case_sensitive: 大文字小文字を区別するか（デフォルト: false）
            max_hits: 最大ヒット数（デフォルト: 50）
    """
    file_path = args.get("file_path", "")
    query = args.get("query", "")
    case_sensitive = args.get("case_sensitive", False)
    max_hits = args.get("max_hits", 50)

    # 古い形式チェック
    old_format_error = _check_old_format(file_path)
    if old_format_error:
        return old_format_error

    if not query:
        return {
            "content": [{"type": "text", "text": "エラー: query（検索キーワード）を指定してください。"}],
            "is_error": True,
        }

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return {
            "content": [{"type": "text", "text": "エラー: openpyxlライブラリがインストールされていません。"}],
            "is_error": True,
        }

    try:
        content, _, _ = await workspace_service.download_file(
            tenant_id, conversation_id, file_path
        )

        result = search_workbook(
            content,
            query,
            case_sensitive=case_sensitive,
            max_hits=max_hits,
        )

        # 結果をテキスト形式でフォーマット
        result_lines = [
            f"# 検索結果: \"{result['query']}\"",
            f"ヒット数: {result['total_hits']}",
            "",
        ]

        if result['hits']:
            for hit in result['hits']:
                result_lines.append(f"## {hit['sheet']}!{hit['cell']}")
                result_lines.append(f"値: {hit['value']}")
                result_lines.append(f"コンテキスト: {hit['context']}")
                result_lines.append("")
        else:
            result_lines.append("検索結果はありませんでした。")

        return {
            "content": [{"type": "text", "text": "\n".join(result_lines)}],
        }
    except FileNotFoundError:
        return {
            "content": [{"type": "text", "text": f"ファイルが見つかりません: {file_path}"}],
            "is_error": True,
        }
    except ValueError as e:
        return {
            "content": [{"type": "text", "text": str(e)}],
            "is_error": True,
        }
    except Exception as e:
        logger.error("Excel検索エラー", error=str(e), file_path=file_path, query=query)
        return {
            "content": [{"type": "text", "text": f"検索エラー: {str(e)}"}],
            "is_error": True,
        }
